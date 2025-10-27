#!/usr/bin/env python3
"""
Parser para Formularios Fehaciente en formato XLSX/XLS.

Este parser extrae datos estructurados de los formularios Fehaciente
(Solicitud de Uso de Capacidad Técnica Dedicada) en formato Excel.

Estructura del archivo XLSX:
- Labels en columna B (col 2)
- Valores en columnas D, E, F, G, H (col 4-8) dependiendo del campo
- ~71 filas × 31 columnas
- Versión template: 2504-FORM-Fehaciente-V1

Fecha: 2025-10-20
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class FehacienteXLSXParser:
    """Parser para Formularios Fehaciente en formato XLSX/XLS."""

    VERSION = "1.0.0"

    def __init__(self):
        """Inicializa el parser."""
        self.logger = logger

    def parse(self, xlsx_path: str) -> Dict[str, Any]:
        """
        Parsea un archivo XLSX de Formulario Fehaciente.

        Args:
            xlsx_path: Ruta al archivo XLSX

        Returns:
            Diccionario con los datos extraídos del formulario

        Raises:
            FileNotFoundError: Si el archivo no existe
            Exception: Si hay error en el parsing
        """
        xlsx_file = Path(xlsx_path)

        if not xlsx_file.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {xlsx_path}")

        self.logger.info(f"📊 Parseando XLSX Fehaciente: {xlsx_file.name}")

        try:
            # Abrir workbook
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)

            # Buscar hoja "FORMULARIO PROYECTO FEHACIENTE" (puede tener variaciones)
            sheet = None
            for sheet_name in wb.sheetnames:
                if 'FORMULARIO' in sheet_name.upper() and 'FEHACIENTE' in sheet_name.upper():
                    sheet = wb[sheet_name]
                    self.logger.debug(f"📄 Usando hoja: {sheet_name}")
                    break

            # Si no encuentra la hoja específica, usar la primera
            if sheet is None:
                sheet = wb.worksheets[0]
                self.logger.warning(f"⚠️  No se encontró hoja 'FORMULARIO PROYECTO FEHACIENTE', usando primera hoja: {sheet.title}")

            self.logger.debug(
                f"📄 Sheet: {sheet.title} | Dimensiones: {sheet.max_row}x{sheet.max_column}"
            )

            # Parsear datos del worksheet
            data = self._parse_worksheet(sheet)

            # Cerrar workbook
            wb.close()

            self.logger.info(f"✅ Parsing XLSX exitoso: {len(data)} campos extraídos")

            return data

        except Exception as e:
            self.logger.error(f"❌ Error parseando XLSX Fehaciente: {e}", exc_info=True)
            raise

    def _parse_worksheet(self, sheet: Worksheet) -> Dict[str, Any]:
        """
        Extrae datos del worksheet.

        Estructura del XLSX Fehaciente (basada en docs reales, no en template):
        - Labels siempre en columna B (col 2)
        - Valores varían entre columnas D, E, F, G, H (cols 4-8)

        Mapeo de filas clave (DOCS REALES):
        F6:  Razón Social → D
        F7:  RUT → D
        F9:  Domicilio Legal → D
        F11: Nombre Rep Legal → F
        F12: Email Rep Legal → D, Teléfono → H
        F14: Coordinador 1 Nombre → F
        F15: Coordinador 1 Email → D, Teléfono → H
        F16: Coordinador 2 Nombre → F
        F17: Coordinador 2 Email → D, Teléfono → H
        F21: Nombre Proyecto → E ← ¡CRÍTICO!
        F22: Tipo Proyecto → G, Potencia Nominal → H
        F23: Consumo propio → G, Factor Potencia → H
        F25: Tipo Tecnología → B (¡en col B!)
        F27: Coordenadas proyecto: Huso (F), Este (H)
        F28: Comuna proyecto → D, Región → H
        F32: Tipo Conexión → H
        F33: Nombre S/E o Línea → H
        F34: Distancia seccionamiento → H
        F35: Nivel Tensión → H
        F36: Carácter conexión → H
        F37: Fecha construcción → H
        F38: Fecha operación → H
        F40: Coordenadas conexión: Huso (F), Este (H)
        F41: Comuna conexión → D, Región → H

        NOTA: Los docs reales NO contienen campos de BESS (almacenamiento),
        modo_control_inversores, ni algunos otros del schema.
        Estos campos se dejan como None para compatibilidad.

        Args:
            sheet: Worksheet de openpyxl

        Returns:
            Diccionario con datos normalizados
        """
        data = {}

        # Helper para obtener valor de celda
        def get_cell(row: int, col: int, as_string: bool = True) -> Optional[Any]:
            """
            Obtiene valor de celda.

            Args:
                row: Número de fila
                col: Número de columna
                as_string: Si True, convierte a string. Si False, retorna raw value.

            Returns:
                Valor de la celda (string o raw)
            """
            value = sheet.cell(row=row, column=col).value
            if value is None:
                return None

            if as_string:
                return str(value).strip()
            else:
                return value

        # === SECCIÓN 1: ANTECEDENTES DE LA EMPRESA SOLICITANTE ===

        # Razón Social (F6, D=col 4)
        data["razon_social"] = get_cell(6, 4)

        # RUT (F7, D=col 4)
        rut_raw = get_cell(7, 4)
        data["rut"] = self._normalize_rut(rut_raw)

        # Domicilio Legal (F9, D=col 4)
        data["domicilio_legal"] = get_cell(9, 4)

        # === CONTACTO DE REPRESENTANTE LEGAL ===

        # Nombre (F11, F=col 6)
        data["representante_legal_nombre"] = get_cell(11, 6)

        # Email (F12, D=col 4)
        data["representante_legal_email"] = get_cell(12, 4)

        # Teléfono (F12, H=col 8) - mismo row que email
        data["representante_legal_telefono"] = get_cell(12, 8)

        # === COORDINADORES DE PROYECTOS ===

        # Primer coordinador (F14, F=col 6)
        data["coordinador_proyecto_1_nombre"] = get_cell(14, 6)
        data["coordinador_proyecto_1_email"] = get_cell(15, 4)
        data["coordinador_proyecto_1_telefono"] = get_cell(15, 8)

        # Segundo coordinador (F16, F=col 6)
        data["coordinador_proyecto_2_nombre"] = get_cell(16, 6)
        data["coordinador_proyecto_2_email"] = get_cell(17, 4)
        data["coordinador_proyecto_2_telefono"] = get_cell(17, 8)

        # === SECCIÓN 2: ANTECEDENTES DEL PROYECTO ===

        # Nombre del Proyecto (F21, E=col 5) ← ¡CRÍTICO!
        data["nombre_proyecto"] = get_cell(21, 5)

        # Tipo de Proyecto (F22, G=col 7)
        data["tipo_proyecto"] = get_cell(22, 7)

        # Potencia Nominal [MW] (F22, H=col 8) - mismo row que tipo_proyecto
        potencia_nominal_raw = get_cell(22, 8)
        data["potencia_neta_inyeccion_mw"] = self._parse_float(potencia_nominal_raw)

        # Consumo propio [MW] (F23, G=col 7)
        consumo_propio_raw = get_cell(23, 7)
        # data["consumo_propio_mw"] = self._parse_float(consumo_propio_raw)  # No existe en schema

        # Factor de Potencia (F23, H=col 8)
        factor_potencia_raw = get_cell(23, 8)
        data["factor_potencia_nominal"] = self._parse_float(factor_potencia_raw)

        # Tipo de Tecnología (F25, B=col 2) - ¡En col B, no D!
        data["tipo_tecnologia"] = get_cell(25, 2)

        # Modo de operación inversores - campo no visible en docs reales
        data["modo_control_inversores"] = None

        # Potencia neta retiro - campo no visible en docs reales
        data["potencia_neta_retiro_mw"] = None

        # === PARÁMETROS SISTEMAS DE ALMACENAMIENTO DE ENERGÍA ===
        # Nota: Estos campos NO están presentes en los docs reales descargados
        # Se dejan como None para mantener compatibilidad con schema

        data["componente_generacion"] = None
        data["componente_generacion_potencia_mw"] = None
        data["componente_almacenamiento"] = None
        data["componente_almacenamiento_potencia_mw"] = None
        data["componente_almacenamiento_energia_mwh"] = None
        data["componente_almacenamiento_horas"] = None

        # === SECCIÓN 3: UBICACIÓN GEOGRÁFICA DEL PROYECTO ===

        # Coordenadas UTM WGS84 (F27)
        # Estructura: F=Huso, H=Este, I=Norte (probablemente)
        data["proyecto_coordenadas_utm_huso"] = get_cell(27, 6)  # F27

        # Este (H27)
        este_raw = get_cell(27, 8)
        data["proyecto_coordenadas_utm_este"] = self._parse_coordinate(este_raw)

        # Norte (I27) - Nota: en docs reales no está visible, buscar en filas siguientes
        # Por ahora dejarlo vacío
        norte_raw = None
        data["proyecto_coordenadas_utm_norte"] = self._parse_coordinate(norte_raw)

        # Comuna (F28, D=col 4)
        data["proyecto_comuna"] = get_cell(28, 4)

        # Región (F28, H=col 8)
        data["proyecto_region"] = get_cell(28, 8)

        # === SECCIÓN 4: ANTECEDENTES DEL PUNTO DE CONEXIÓN ===

        # Tipo de Conexión (F32, H=col 8)
        data["tipo_conexion"] = get_cell(32, 8)

        # Nombre de la S/E o Línea de Transmisión (F33, H=col 8)
        data["nombre_se_o_linea"] = get_cell(33, 8)

        # Distancia seccionamiento/derivación (F34, H=col 8)
        distancia_raw = get_cell(34, 8)
        data["seccionamiento_distancia_km"] = self._parse_float(distancia_raw)

        # Nivel de Tensión [kV] (F35, H=col 8)
        data["nivel_tension_kv"] = get_cell(35, 8)

        # Carácter de conexión (F36, H=col 8)
        caracter_conexion = get_cell(36, 8)
        # data["caracter_conexion"] = caracter_conexion  # No existe en schema

        # Fecha estimada de Declaración en Construcción (F37, H=col 8)
        fecha_construccion_raw = get_cell(37, 8, as_string=False)
        data["fecha_estimada_construccion"] = self._parse_date(fecha_construccion_raw)

        # Fecha estimada de Interconexión (F38, H=col 8)
        fecha_operacion_raw = get_cell(38, 8, as_string=False)
        data["fecha_estimada_operacion"] = self._parse_date(fecha_operacion_raw)

        # S/E cercana a seccionamiento - campo no visible en docs reales
        data["seccionamiento_se_cercana"] = None

        # Paño o N° de estructura(s) - campo no visible en docs reales
        data["pano_o_estructura"] = None

        # === SECCIÓN 5: UBICACIÓN GEOGRÁFICA DEL PUNTO DE CONEXIÓN ===

        # Coordenadas UTM WGS84 punto de conexión (F40)
        # Estructura: F=Huso, H=Este, I=Norte (probablemente)
        data["conexion_coordenadas_utm_huso"] = get_cell(40, 6)  # F40

        # Este (H40)
        conexion_este_raw = get_cell(40, 8)
        data["conexion_coordenadas_utm_este"] = self._parse_coordinate(conexion_este_raw)

        # Norte (I40) - No visible en docs reales
        conexion_norte_raw = None
        data["conexion_coordenadas_utm_norte"] = self._parse_coordinate(conexion_norte_raw)

        # Comuna punto conexión (F41, D=col 4)
        data["conexion_comuna"] = get_cell(41, 4)

        # Región punto conexión (F41, H=col 8)
        data["conexion_region"] = get_cell(41, 8)

        # === INFORMACIÓN ADICIONAL (OPCIONAL) ===
        # Puede estar en filas posteriores, por ahora dejamos None
        data["informacion_adicional"] = None

        # =====================================================================
        # NORMALIZACIÓN DE CAMPOS PARA COMPATIBILIDAD CON PDF PARSER
        # =====================================================================
        # Retornamos el diccionario normalizado directamente
        # (ya están con los nombres correctos de BD)

        normalized = data.copy()

        # Metadata del archivo (No aplica para XLSX, solo PDF tiene metadata)
        normalized["pdf_producer"] = None
        normalized["pdf_author"] = None
        normalized["pdf_title"] = None
        normalized["pdf_creation_date"] = None

        return normalized

    # =========================================================================
    # MÉTODOS HELPER DE NORMALIZACIÓN
    # =========================================================================

    def _normalize_rut(self, rut: Optional[str]) -> Optional[str]:
        """
        Normaliza un RUT al formato estándar XX.XXX.XXX-X.

        Args:
            rut: RUT en cualquier formato

        Returns:
            RUT normalizado o None
        """
        if not rut:
            return None

        # Eliminar puntos, guiones, espacios
        rut_clean = rut.replace(".", "").replace("-", "").replace(" ", "").upper()

        if not rut_clean:
            return None

        # Separar cuerpo y dígito verificador
        if len(rut_clean) < 2:
            return rut  # No se puede normalizar

        cuerpo = rut_clean[:-1]
        dv = rut_clean[-1]

        # Formatear: XX.XXX.XXX-X
        if not cuerpo.isdigit():
            return rut  # Formato inválido

        # Agregar puntos cada 3 dígitos desde la derecha
        cuerpo_formateado = ""
        for i, digit in enumerate(reversed(cuerpo)):
            if i > 0 and i % 3 == 0:
                cuerpo_formateado = "." + cuerpo_formateado
            cuerpo_formateado = digit + cuerpo_formateado

        return f"{cuerpo_formateado}-{dv}"

    def _parse_coordinate(self, coord: Optional[str]) -> Optional[float]:
        """
        Convierte coordenada UTM a float.

        Args:
            coord: Coordenada (Este o Norte)

        Returns:
            Coordenada como float o None
        """
        return self._parse_float(coord)

    def _parse_float(self, value: Optional[str]) -> Optional[float]:
        """
        Convierte string a float, manejando separadores decimales.

        Args:
            value: Valor como string

        Returns:
            Float o None
        """
        if not value:
            return None

        try:
            # Eliminar espacios
            value_clean = str(value).strip()

            # Reemplazar coma decimal por punto
            value_clean = value_clean.replace(",", ".")

            # Eliminar puntos de miles (si hay más de un punto)
            if value_clean.count(".") > 1:
                # Ejemplo: "1.234.567,89" → "1234567.89"
                parts = value_clean.split(".")
                if len(parts) > 2:
                    # Último es decimal, el resto son miles
                    value_clean = "".join(parts[:-1]) + "." + parts[-1]

            return float(value_clean)

        except (ValueError, AttributeError, TypeError):
            self.logger.warning(f"⚠️  No se pudo convertir a float: {value}")
            return None

    def _parse_date(self, date_value: Optional[Any]) -> Optional[str]:
        """
        Convierte fecha a formato MySQL YYYY-MM-DD.

        Formatos soportados:
        - datetime object (from openpyxl)
        - DD-MM-YYYY
        - DD/MM/YYYY
        - YYYY-MM-DD
        - Fecha de Excel (serial number)

        Args:
            date_value: Fecha como datetime, string o número

        Returns:
            Fecha en formato YYYY-MM-DD o None
        """
        if not date_value:
            return None

        try:
            # Si es datetime object (openpyxl ya lo convirtió)
            if isinstance(date_value, datetime):
                return date_value.strftime("%Y-%m-%d")

            # Si es un número (fecha de Excel)
            if isinstance(date_value, (int, float)):
                # Excel serial date (días desde 1900-01-01)
                from datetime import timedelta

                excel_epoch = datetime(1899, 12, 30)  # Excel epoch
                fecha_dt = excel_epoch + timedelta(days=float(date_value))
                return fecha_dt.strftime("%Y-%m-%d")

            # Si es string
            date_str = str(date_value).strip()

            # Probar DD-MM-YYYY o DD/MM/YYYY
            for separator in ["-", "/"]:
                if separator in date_str:
                    parts = date_str.split(separator)
                    if len(parts) == 3:
                        day, month, year = parts
                        # Validar que sea fecha válida
                        fecha_dt = datetime(
                            int(year), int(month), int(day)
                        )  # Lanza ValueError si inválido
                        return fecha_dt.strftime("%Y-%m-%d")

            # Probar YYYY-MM-DD
            if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
                fecha_dt = datetime.strptime(date_str, "%Y-%m-%d")
                return fecha_dt.strftime("%Y-%m-%d")

            self.logger.warning(f"⚠️  Formato de fecha no reconocido: {date_value}")
            return None

        except (ValueError, TypeError) as e:
            self.logger.warning(f"⚠️  Fecha inválida: {date_value} - {e}")
            return None


# ============================================================================
# FUNCIÓN PRINCIPAL PARA TESTING
# ============================================================================


def main():
    """Función de prueba del parser."""
    import sys

    if len(sys.argv) < 2:
        print("Uso: python -m src.parsers.xlsx_fehaciente <ruta_xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]

    # Configurar logging
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
    )

    # Parsear
    parser = FehacienteXLSXParser()
    try:
        data = parser.parse(xlsx_path)

        print("\n" + "=" * 70)
        print("📊 DATOS EXTRAÍDOS DEL FORMULARIO Fehaciente (XLSX)")
        print("=" * 70)

        for key, value in data.items():
            if value is not None:
                print(f"{key:40s}: {value}")

        print("=" * 70)

    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
