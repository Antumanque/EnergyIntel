#!/usr/bin/env python3
"""
Parser para Formularios SAC en formato XLSX/XLS.

Este parser extrae datos estructurados de los formularios SAC (Solicitud de Acceso y Conexión)
en formato Excel, específicamente para documentos generados desde plantillas XLSX.

Estructura del archivo XLSX:
- Labels en columna B (col 2)
- Valores en columnas D, E, F, G (col 4-7) dependiendo del campo
- ~49 filas × 12 columnas
- Estructura más predecible que PDF

Fecha: 2025-10-20
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class SACXLSXParser:
    """Parser para Formularios SAC en formato XLSX/XLS."""

    VERSION = "1.0.0"

    def __init__(self):
        """Inicializa el parser."""
        self.logger = logger

    def parse(self, xlsx_path: str) -> Dict[str, Any]:
        """
        Parsea un archivo XLSX de Formulario SAC.

        Args:
            xlsx_path: Ruta al archivo XLSX

        Returns:
            Diccionario con los datos extraídos del formulario

        Raises:
            FileNotFoundError: Si el archivo no existe
            Exception: Si hay error en el parsing
        """
        xlsx_file = Path(xlsx_path)

        if not xlsx_file.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {xlsx_path}")

        self.logger.info(f"📊 Parseando XLSX SAC: {xlsx_file.name}")

        try:
            # Abrir workbook
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            sheet = wb.active

            self.logger.debug(
                f"📄 Sheet: {sheet.title} | Dimensiones: {sheet.max_row}x{sheet.max_column}"
            )

            # Parsear datos del worksheet
            data = self._parse_worksheet(sheet)

            # Cerrar workbook
            wb.close()

            self.logger.info(f"✅ Parsing XLSX exitoso: {len(data)} campos extraídos")

            return data

        except Exception as e:
            self.logger.error(f"❌ Error parseando XLSX SAC: {e}", exc_info=True)
            raise

    def _parse_worksheet(self, sheet: Worksheet) -> Dict[str, Any]:
        """
        Extrae datos del worksheet.

        Estructura descubierta del XLSX real:
        - Labels siempre en columna B (col 2)
        - Valores varían entre columnas D, E, F, G, H (cols 4-8)

        Mapeo correcto:
        F 6: B=Razón Social | D=valor
        F 7: B=RUT | D=valor
        F11: B=Rep Legal Nombre | F=valor
        F12: B=Email | D=valor
        F21: B=Nombre Proyecto | E=valor
        F27: B=Coordenadas | E=Huso, F=Zona, G=Este, H=Norte

        Args:
            sheet: Worksheet de openpyxl

        Returns:
            Diccionario con datos normalizados
        """
        data = {}

        # Helper para obtener valor de celda
        def get_cell(row: int, col: int, as_string: bool = True) -> Optional[Any]:
            """
            Obtiene valor de celda.

            Args:
                row: Número de fila
                col: Número de columna
                as_string: Si True, convierte a string. Si False, retorna raw value.

            Returns:
                Valor de la celda (string o raw)
            """
            value = sheet.cell(row=row, column=col).value
            if value is None:
                return None

            if as_string:
                return str(value).strip()
            else:
                return value

        # === SECCIÓN 1: INFORMACIÓN DEL SOLICITANTE ===

        # Razón Social (F6, D=col 4)
        data["razon_social"] = get_cell(6, 4)

        # RUT (F7, D=col 4)
        rut_raw = get_cell(7, 4)
        data["rut"] = self._normalize_rut(rut_raw)

        # Giro (F8, D=col 4)
        data["giro"] = get_cell(8, 4)

        # Domicilio Legal (F9, D=col 4)
        data["domicilio"] = get_cell(9, 4)

        # Representante Legal Nombre (F11, F=col 6)
        data["representante_legal_nombre"] = get_cell(11, 6)

        # Representante Legal Email (F12, D=col 4)
        data["representante_legal_email"] = get_cell(12, 4)

        # Representante Legal Teléfono (F12, H=col 8) - valor después de "Teléfono"
        # Nota: El label "Teléfono" está en H12, valor probablemente en siguiente celda
        # Por ahora dejamos None, necesitaríamos ver estructura exacta

        # Coordinadores de proyecto
        # Primer coordinador (F14, F=col 6)
        data["coordinador1_nombre"] = get_cell(14, 6)
        data["coordinador1_email"] = get_cell(15, 4)  # F15, D

        # Segundo coordinador (F16, F=col 6)
        data["coordinador2_nombre"] = get_cell(16, 6)
        data["coordinador2_email"] = get_cell(17, 4)  # F17, D

        # === SECCIÓN 2: IDENTIFICACIÓN DEL PROYECTO ===

        # Nombre del Proyecto (F21, E=col 5)
        data["proyecto_nombre"] = get_cell(21, 5)

        # Tipo de Proyecto (F22, G=col 7)
        data["proyecto_tipo"] = get_cell(22, 7)

        # Potencia Nominal (F22, H=col 8)
        potencia_nominal_raw = get_cell(22, 8)
        data["potencia_nominal"] = self._parse_float(potencia_nominal_raw)

        # Consumo Propio (F23, G=col 7)
        consumo_propio_raw = get_cell(23, 7)
        data["consumo_propio"] = self._parse_float(consumo_propio_raw)

        # Factor de Potencia (F23, H=col 8)
        factor_potencia_raw = get_cell(23, 8)
        data["factor_potencia"] = self._parse_float(factor_potencia_raw)

        # === SECCIÓN 3: UBICACIÓN GEOGRÁFICA DEL PROYECTO ===

        # Coordenadas UTM WGS84 (F27)
        # Estructura: E=Label(Huso), F=Valor(18H), G=Label(Este), H=Valor(662975)
        # Norte no siempre está presente (merged cells o fila adicional)
        data["utm_huso_label"] = get_cell(27, 5)  # "Huso"
        data["utm_huso"] = get_cell(27, 6)  # "18H"

        # Este: label en G, valor en H
        este_raw = get_cell(27, 8)  # H27 (not G27)
        data["utm_este"] = self._parse_coordinate(este_raw)

        # Norte: Puede estar en I27 o en siguiente fila
        # Por ahora dejamos None si no está presente
        norte_raw = get_cell(27, 9)  # I27
        if norte_raw:
            data["utm_norte"] = self._parse_coordinate(norte_raw)
        else:
            data["utm_norte"] = None

        # Comuna (F28, D=col 4)
        data["proyecto_comuna"] = get_cell(28, 4)

        # Región (F28, H=col 8)
        data["proyecto_region"] = get_cell(28, 8)

        # === SECCIÓN 4: PUNTO DE CONEXIÓN ===

        # Nombre de S/E (F32, H=col 8)
        data["subestacion_nombre"] = get_cell(32, 8)

        # Nivel de Tensión (F33, H=col 8)
        data["tension_conexion"] = get_cell(33, 8)

        # Carácter de conexión (F34, H=col 8)
        data["caracter_conexion"] = get_cell(34, 8)

        # Fecha estimada Declaración Construcción (F35, H=col 8)
        # Usar as_string=False para mantener datetime object
        fecha_construccion_raw = get_cell(35, 8, as_string=False)
        data["fecha_estimada_construccion"] = self._parse_date(fecha_construccion_raw)

        # Fecha estimada Interconexión (F36, H=col 8)
        fecha_interconexion_raw = get_cell(36, 8, as_string=False)
        data["fecha_estimada_interconexion"] = self._parse_date(fecha_interconexion_raw)

        # === UBICACIÓN PUNTO DE CONEXIÓN ===

        # Coordenadas punto de conexión (F38)
        # Misma estructura que proyecto: E=Label, F=Valor, G=Label, H=Valor
        data["conexion_utm_huso"] = get_cell(38, 6)  # F38 (not E38)

        # Este: valor en H38
        conexion_este_raw = get_cell(38, 8)  # H38
        data["conexion_utm_este"] = self._parse_coordinate(conexion_este_raw)

        # Norte: en I38 si existe
        conexion_norte_raw = get_cell(38, 9)  # I38
        if conexion_norte_raw:
            data["conexion_utm_norte"] = self._parse_coordinate(conexion_norte_raw)
        else:
            data["conexion_utm_norte"] = None

        # Comuna punto conexión (F39, D=col 4)
        data["conexion_comuna"] = get_cell(39, 4)

        # Región punto conexión (F39, H=col 8)
        data["conexion_region"] = get_cell(39, 8)

        # =====================================================================
        # NORMALIZACIÓN DE CAMPOS PARA COMPATIBILIDAD CON PDF PARSER
        # =====================================================================
        # El PDF parser usa nombres ligeramente diferentes.
        # Aquí mapeamos los campos del XLSX a los nombres esperados por el repositorio.

        normalized = {}

        # Sección 1: Solicitante
        normalized["razon_social"] = data.get("razon_social")
        normalized["rut"] = data.get("rut")
        normalized["giro"] = data.get("giro")
        normalized["domicilio_legal"] = data.get("domicilio")

        normalized["representante_legal_nombre"] = data.get("representante_legal_nombre")
        normalized["representante_legal_email"] = data.get("representante_legal_email")
        normalized["representante_legal_telefono"] = None  # No presente en XLSX actual

        # Coordinadores
        normalized["coordinador_proyecto_1_nombre"] = data.get("coordinador1_nombre")
        normalized["coordinador_proyecto_1_email"] = data.get("coordinador1_email")
        normalized["coordinador_proyecto_1_telefono"] = None

        normalized["coordinador_proyecto_2_nombre"] = data.get("coordinador2_nombre")
        normalized["coordinador_proyecto_2_email"] = data.get("coordinador2_email")
        normalized["coordinador_proyecto_2_telefono"] = None

        normalized["coordinador_proyecto_3_nombre"] = None
        normalized["coordinador_proyecto_3_email"] = None
        normalized["coordinador_proyecto_3_telefono"] = None

        # Sección 2: Proyecto
        normalized["nombre_proyecto"] = data.get("proyecto_nombre")
        normalized["tipo_proyecto"] = data.get("proyecto_tipo")
        normalized["tecnologia"] = None  # No presente en XLSX actual
        normalized["potencia_nominal_mw"] = data.get("potencia_nominal")
        normalized["consumo_propio_mw"] = data.get("consumo_propio")
        normalized["factor_potencia"] = data.get("factor_potencia")

        # Sección 3: Ubicación proyecto
        normalized["proyecto_coordenadas_utm_huso"] = data.get("utm_huso")
        normalized["proyecto_coordenadas_utm_este"] = data.get("utm_este")
        normalized["proyecto_coordenadas_utm_norte"] = data.get("utm_norte")
        normalized["proyecto_comuna"] = data.get("proyecto_comuna")
        normalized["proyecto_region"] = data.get("proyecto_region")

        # Sección 4: Punto de conexión
        normalized["nombre_subestacion"] = data.get("subestacion_nombre")
        normalized["nivel_tension_kv"] = data.get("tension_conexion")
        normalized["caracter_conexion"] = data.get("caracter_conexion")
        normalized["fecha_estimada_construccion"] = data.get("fecha_estimada_construccion")
        normalized["fecha_estimada_interconexion"] = data.get("fecha_estimada_interconexion")

        # Ubicación punto conexión
        normalized["conexion_coordenadas_utm_huso"] = data.get("conexion_utm_huso")
        normalized["conexion_coordenadas_utm_este"] = data.get("conexion_utm_este")
        normalized["conexion_coordenadas_utm_norte"] = data.get("conexion_utm_norte")
        normalized["conexion_comuna"] = data.get("conexion_comuna")
        normalized["conexion_region"] = data.get("conexion_region")

        # Metadata del archivo (No aplica para XLSX, solo PDF tiene metadata)
        normalized["pdf_producer"] = None
        normalized["pdf_author"] = None
        normalized["pdf_title"] = None
        normalized["pdf_creation_date"] = None

        return normalized

    # =========================================================================
    # MÉTODOS HELPER DE NORMALIZACIÓN
    # =========================================================================

    def _normalize_rut(self, rut: Optional[str]) -> Optional[str]:
        """
        Normaliza un RUT al formato estándar XX.XXX.XXX-X.

        Args:
            rut: RUT en cualquier formato

        Returns:
            RUT normalizado o None
        """
        if not rut:
            return None

        # Eliminar puntos, guiones, espacios
        rut_clean = rut.replace(".", "").replace("-", "").replace(" ", "").upper()

        if not rut_clean:
            return None

        # Separar cuerpo y dígito verificador
        if len(rut_clean) < 2:
            return rut  # No se puede normalizar

        cuerpo = rut_clean[:-1]
        dv = rut_clean[-1]

        # Formatear: XX.XXX.XXX-X
        if not cuerpo.isdigit():
            return rut  # Formato inválido

        # Agregar puntos cada 3 dígitos desde la derecha
        cuerpo_formateado = ""
        for i, digit in enumerate(reversed(cuerpo)):
            if i > 0 and i % 3 == 0:
                cuerpo_formateado = "." + cuerpo_formateado
            cuerpo_formateado = digit + cuerpo_formateado

        return f"{cuerpo_formateado}-{dv}"

    def _parse_potencia(
        self, potencia: Optional[str], unidad: Optional[str]
    ) -> Optional[float]:
        """
        Convierte potencia a float.

        Args:
            potencia: Valor de potencia (puede incluir separadores)
            unidad: Unidad (MW, kW, etc.)

        Returns:
            Potencia como float o None
        """
        return self._parse_float(potencia)

    def _parse_tension(
        self, tension: Optional[str], unidad: Optional[str]
    ) -> Optional[float]:
        """
        Convierte tensión a float.

        Args:
            tension: Valor de tensión
            unidad: Unidad (kV, etc.)

        Returns:
            Tensión como float o None
        """
        return self._parse_float(tension)

    def _parse_coordinate(self, coord: Optional[str]) -> Optional[float]:
        """
        Convierte coordenada UTM a float.

        Args:
            coord: Coordenada (Este o Norte)

        Returns:
            Coordenada como float o None
        """
        return self._parse_float(coord)

    def _parse_float(self, value: Optional[str]) -> Optional[float]:
        """
        Convierte string a float, manejando separadores decimales.

        Args:
            value: Valor como string

        Returns:
            Float o None
        """
        if not value:
            return None

        try:
            # Eliminar espacios
            value_clean = value.strip()

            # Reemplazar coma decimal por punto
            value_clean = value_clean.replace(",", ".")

            # Eliminar puntos de miles (si hay más de un punto)
            if value_clean.count(".") > 1:
                # Ejemplo: "1.234.567,89" → "1234567.89"
                parts = value_clean.split(".")
                if len(parts) > 2:
                    # Último es decimal, el resto son miles
                    value_clean = "".join(parts[:-1]) + "." + parts[-1]

            return float(value_clean)

        except (ValueError, AttributeError):
            self.logger.warning(f"⚠️  No se pudo convertir a float: {value}")
            return None

    def _parse_date(self, date_value: Optional[Any]) -> Optional[str]:
        """
        Convierte fecha a formato MySQL YYYY-MM-DD.

        Formatos soportados:
        - datetime object (from openpyxl)
        - DD-MM-YYYY
        - DD/MM/YYYY
        - YYYY-MM-DD
        - Fecha de Excel (serial number)

        Args:
            date_value: Fecha como datetime, string o número

        Returns:
            Fecha en formato YYYY-MM-DD o None
        """
        if not date_value:
            return None

        try:
            # Si es datetime object (openpyxl ya lo convirtió)
            if isinstance(date_value, datetime):
                return date_value.strftime("%Y-%m-%d")

            # Si es un número (fecha de Excel)
            if isinstance(date_value, (int, float)):
                # Excel serial date (días desde 1900-01-01)
                from datetime import timedelta

                excel_epoch = datetime(1899, 12, 30)  # Excel epoch
                fecha_dt = excel_epoch + timedelta(days=float(date_value))
                return fecha_dt.strftime("%Y-%m-%d")

            # Si es string
            date_str = str(date_value).strip()

            # Probar DD-MM-YYYY o DD/MM/YYYY
            for separator in ["-", "/"]:
                if separator in date_str:
                    parts = date_str.split(separator)
                    if len(parts) == 3:
                        day, month, year = parts
                        # Validar que sea fecha válida
                        fecha_dt = datetime(
                            int(year), int(month), int(day)
                        )  # Lanza ValueError si inválido
                        return fecha_dt.strftime("%Y-%m-%d")

            # Probar YYYY-MM-DD
            if len(date_str) == 10 and date_str[4] == "-" and date_str[7] == "-":
                fecha_dt = datetime.strptime(date_str, "%Y-%m-%d")
                return fecha_dt.strftime("%Y-%m-%d")

            self.logger.warning(f"⚠️  Formato de fecha no reconocido: {date_value}")
            return None

        except (ValueError, TypeError) as e:
            self.logger.warning(f"⚠️  Fecha inválida: {date_value} - {e}")
            return None


# ============================================================================
# FUNCIÓN PRINCIPAL PARA TESTING
# ============================================================================


def main():
    """Función de prueba del parser."""
    import sys

    if len(sys.argv) < 2:
        print("Uso: python -m src.parsers.xlsx_sac <ruta_xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]

    # Configurar logging
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
    )

    # Parsear
    parser = SACXLSXParser()
    try:
        data = parser.parse(xlsx_path)

        print("\n" + "=" * 70)
        print("📊 DATOS EXTRAÍDOS DEL FORMULARIO SAC (XLSX)")
        print("=" * 70)

        for key, value in data.items():
            if value is not None:
                print(f"{key:30s}: {value}")

        print("=" * 70)

    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
