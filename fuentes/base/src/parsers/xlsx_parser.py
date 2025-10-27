"""
Parser para archivos Excel (XLSX).

Este módulo provee funcionalidad básica para leer datos desde archivos Excel.
"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.parsers.base import BaseParser

logger = logging.getLogger(__name__)


class XLSXParser(BaseParser):
    """
    Parser para archivos Excel usando openpyxl.

    Este parser lee hojas de Excel y extrae los datos.
    Para casos de uso específicos, hereda de esta clase y personaliza.
    """

    def __init__(self, sheet_name: str | None = None, data_only: bool = True):
        """
        Inicializar el parser XLSX.

        Args:
            sheet_name: Nombre de hoja específica a leer (None = hoja activa)
            data_only: Si True, lee valores en lugar de fórmulas
        """
        super().__init__()
        self.sheet_name = sheet_name
        self.data_only = data_only

    def parse(self, data: Any) -> dict[str, Any]:
        """
        Parsear archivo Excel.

        Args:
            data: Ruta del archivo XLSX (str o Path)

        Returns:
            Diccionario con resultado del parseo
        """
        try:
            file_path = Path(data) if not isinstance(data, Path) else data

            if not file_path.exists():
                raise FileNotFoundError(f"XLSX file not found: {file_path}")

            workbook = load_workbook(file_path, data_only=self.data_only)

            # Get the sheet to parse
            if self.sheet_name:
                sheet = workbook[self.sheet_name]
            else:
                sheet = workbook.active

            # Extract all rows
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(list(row))

            # Extract sheet names
            sheet_names = workbook.sheetnames

            parsed_data = {
                "sheet_name": sheet.title,
                "rows": rows,
                "num_rows": len(rows),
                "num_cols": len(rows[0]) if rows else 0,
                "all_sheets": sheet_names,
            }

            workbook.close()

            return {
                "parsing_successful": True,
                "parsed_data": parsed_data,
                "error_message": None,
                "metadata": {
                    "parser_type": "xlsx",
                    "num_sheets": len(sheet_names),
                    "num_rows": len(rows),
                },
            }

        except FileNotFoundError as e:
            logger.error(f"XLSX file not found: {e}")
            return {
                "parsing_successful": False,
                "parsed_data": None,
                "error_message": str(e),
                "metadata": {"parser_type": "xlsx"},
            }

        except Exception as e:
            logger.error(f"Error parsing XLSX: {e}", exc_info=True)
            return {
                "parsing_successful": False,
                "parsed_data": None,
                "error_message": f"XLSX parsing error: {str(e)}",
                "metadata": {"parser_type": "xlsx"},
            }
